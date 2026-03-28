#!/usr/bin/env python3
"""
PDF 转 Excel
用法: pdf_to_excel.py <input.pdf> [output.xlsx]
"""

import sys
import os
import pdfplumber
from openpyxl import Workbook

def pdf_to_excel(pdf_path, excel_path):
    """将 PDF 转换为 Excel"""

    if not os.path.exists(pdf_path):
        print(f"❌ 错误: 文件不存在: {pdf_path}")
        return False

    print(f"📄 转换 PDF → Excel")
    print(f"   输入: {pdf_path}")
    print(f"   输出: {excel_path}")
    print()

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "提取内容"

        row_num = 1

        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            print(f"   总页数: {total_pages}")

            for page_num, page in enumerate(pdf.pages, 1):
                print(f"   处理第 {page_num}/{total_pages} 页...", end='\r')

                # 提取文本
                text = page.extract_text()
                if text:
                    ws.cell(row=row_num, column=1, value=f"=== 第 {page_num} 页 ===")
                    row_num += 1

                    for line in text.split('\n'):
                        ws.cell(row=row_num, column=1, value=line)
                        row_num += 1

                    row_num += 1  # 空行分隔

                # 提取表格
                tables = page.extract_tables()
                if tables:
                    for table_num, table in enumerate(tables, 1):
                        ws.cell(row=row_num, column=1, value=f"--- 表格 {table_num} ---")
                        row_num += 1

                        for row in table:
                            for col_num, cell in enumerate(row, 1):
                                ws.cell(row=row_num, column=col_num, value=str(cell) if cell else "")
                            row_num += 1

                        row_num += 1  # 空行分隔

            print(f"   处理完成！                    ")

        wb.save(excel_path)

        if os.path.exists(excel_path):
            size = os.path.getsize(excel_path)
            print()
            print(f"✅ 转换成功！")
            print(f"   输出文件: {excel_path}")
            print(f"   文件大小: {size / 1024:.2f} KB")
            return True
        else:
            print("❌ 转换失败")
            return False

    except Exception as e:
        print(f"❌ 错误: {e}")
        return False

def main():
    if len(sys.argv) < 2:
        print("用法: python3 pdf_to_excel.py <input.pdf> [output.xlsx]")
        print("示例: python3 pdf_to_excel.py document.pdf")
        print("      python3 pdf_to_excel.py document.pdf output.xlsx")
        sys.exit(1)

    pdf_path = sys.argv[1]

    if len(sys.argv) >= 3:
        excel_path = sys.argv[2]
    else:
        excel_path = pdf_path.replace('.pdf', '.xlsx')

    success = pdf_to_excel(pdf_path, excel_path)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
